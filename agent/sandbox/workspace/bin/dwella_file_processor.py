#!/usr/bin/env python3
import argparse
import hashlib
import json
import mimetypes
import re
import tarfile
import zipfile
from pathlib import Path
from xml.etree import ElementTree


TEXT_EXTENSIONS = {
    ".csv", ".css", ".html", ".htm", ".js", ".json", ".jsx", ".log", ".md",
    ".py", ".text", ".ts", ".tsx", ".txt", ".xml", ".yaml", ".yml",
}


def main():
    parser = argparse.ArgumentParser(description="Inspect and extract uploaded Dwella files in the Eve sandbox.")
    parser.add_argument("--input", default="/workspace/attachments")
    parser.add_argument("--output", default="/workspace/processed-files")
    parser.add_argument("--operation", choices=["inspect", "extract", "redact_copy"], default="extract")
    parser.add_argument("--redact", action="append", default=[])
    args = parser.parse_args()

    input_root = Path(args.input)
    output_root = Path(args.output)
    output_root.mkdir(parents=True, exist_ok=True)
    (output_root / "extracted-text").mkdir(parents=True, exist_ok=True)
    (output_root / "normalized").mkdir(parents=True, exist_ok=True)
    (output_root / "redacted").mkdir(parents=True, exist_ok=True)

    files = [path for path in input_root.rglob("*") if path.is_file()]
    results = []
    for path in files:
        results.append(process_file(path, input_root, output_root, args))

    manifest = {
        "schemaVersion": 1,
        "inputRoot": str(input_root),
        "outputRoot": str(output_root),
        "operation": args.operation,
        "security": {
            "uploadedFilesExecuted": False,
            "networkRequired": False,
            "malwareScan": "not_performed",
            "retentionPolicy": "normalized originals and extracted artifacts are written under /workspace/processed-files; raw attachment staging may be cleaned after processing",
        },
        "fileCount": len(results),
        "files": results,
    }
    manifest_path = output_root / "file-processing-manifest.json"
    manifest_path.write_text(json.dumps(manifest, indent=2, ensure_ascii=False), encoding="utf-8")
    print(json.dumps(manifest, ensure_ascii=False))


def process_file(path, input_root, output_root, args):
    rel = safe_relative(path, input_root)
    raw = path.read_bytes()
    sha = hashlib.sha256(raw).hexdigest()
    media_type = mimetypes.guess_type(path.name)[0] or "application/octet-stream"
    result = {
        "path": str(path),
        "relativePath": rel,
        "filename": path.name,
        "extension": path.suffix.lower(),
        "mediaType": media_type,
        "bytes": len(raw),
        "sha256": sha,
        "kind": "binary",
        "textExtracted": False,
        "originalPreserved": False,
        "warnings": [],
        "outputs": [],
    }

    copy_path = output_root / "normalized" / f"{slugify(rel)}.bin"
    copy_path.write_bytes(raw)
    result["originalPreserved"] = True
    result["outputs"].append(str(copy_path))

    extracted = extract_text(path, raw, result)
    if extracted:
        result["kind"] = "text"
        result["textExtracted"] = True
        result["textPreview"] = extracted[:2000]
        text_path = output_root / "extracted-text" / f"{slugify(rel)}.txt"
        text_path.write_text(extracted, encoding="utf-8")
        result["outputs"].append(str(text_path))

    if args.operation == "redact_copy":
        redacted = redact_text(extracted or decode_text(raw) or "", args.redact)
        redacted_path = output_root / "redacted" / f"{slugify(rel)}.txt"
        redacted_path.write_text(redacted, encoding="utf-8")
        result["outputs"].append(str(redacted_path))

    return result


def extract_text(path, raw, result):
    ext = path.suffix.lower()
    if ext in TEXT_EXTENSIONS or result["mediaType"].startswith("text/"):
        return decode_text(raw)
    if ext == ".pdf":
        return extract_pdf(path, result)
    if ext == ".docx":
        return extract_docx(path, result)
    if ext == ".xlsx":
        return extract_xlsx(path, result)
    if ext == ".pptx":
        return extract_pptx(path, result)
    if ext == ".zip":
        return extract_zip_listing(path, result)
    if ext in {".tar", ".gz", ".tgz"}:
        return extract_tar_listing(path, result)
    if result["mediaType"].startswith("image/"):
        return extract_image_metadata(path, result)
    return ""


def extract_pdf(path, result):
    try:
        from pypdf import PdfReader
        reader = PdfReader(str(path))
        pages = []
        for page in reader.pages:
            pages.append(page.extract_text() or "")
        return "\n\n".join(pages).strip()
    except Exception as error:
        result["warnings"].append(f"pdf_extract_failed: {error}")
        return ""


def extract_docx(path, result):
    try:
        from docx import Document
        doc = Document(str(path))
        return "\n".join(p.text for p in doc.paragraphs if p.text).strip()
    except Exception:
        try:
            with zipfile.ZipFile(path) as archive:
                xml = archive.read("word/document.xml")
            root = ElementTree.fromstring(xml)
            texts = [node.text for node in root.iter() if node.tag.endswith("}t") and node.text]
            return "\n".join(texts).strip()
        except Exception as error:
            result["warnings"].append(f"docx_extract_failed: {error}")
            return ""


def extract_xlsx(path, result):
    try:
        from openpyxl import load_workbook
        workbook = load_workbook(str(path), read_only=True, data_only=True)
        rows = []
        for sheet in workbook.worksheets:
            rows.append(f"# Sheet: {sheet.title}")
            for row in sheet.iter_rows(values_only=True):
                values = ["" if value is None else str(value) for value in row]
                if any(values):
                    rows.append("\t".join(values))
        return "\n".join(rows).strip()
    except Exception as error:
        result["warnings"].append(f"xlsx_extract_failed: {error}")
        return ""


def extract_pptx(path, result):
    try:
        texts = []
        with zipfile.ZipFile(path) as archive:
            for name in sorted(archive.namelist()):
                if name.startswith("ppt/slides/slide") and name.endswith(".xml"):
                    root = ElementTree.fromstring(archive.read(name))
                    slide_text = [node.text for node in root.iter() if node.tag.endswith("}t") and node.text]
                    if slide_text:
                        texts.append("\n".join(slide_text))
        return "\n\n".join(texts).strip()
    except Exception as error:
        result["warnings"].append(f"pptx_extract_failed: {error}")
        return ""


def extract_zip_listing(path, result):
    try:
        with zipfile.ZipFile(path) as archive:
            return "\n".join(archive.namelist())
    except Exception as error:
        result["warnings"].append(f"zip_list_failed: {error}")
        return ""


def extract_tar_listing(path, result):
    try:
        with tarfile.open(path) as archive:
            return "\n".join(archive.getnames())
    except Exception as error:
        result["warnings"].append(f"tar_list_failed: {error}")
        return ""


def extract_image_metadata(path, result):
    try:
        from PIL import Image
        with Image.open(path) as image:
            return f"Image file: {path.name}\nFormat: {image.format}\nSize: {image.width}x{image.height}\nMode: {image.mode}"
    except Exception as error:
        result["warnings"].append(f"image_metadata_failed: {error}")
        return f"Image file: {path.name}"


def decode_text(raw):
    for encoding in ("utf-8", "utf-16", "latin-1"):
        try:
            return raw.decode(encoding).strip()
        except UnicodeDecodeError:
            continue
    return ""


def redact_text(text, terms):
    redacted = text
    for term in [term for term in terms if term]:
        redacted = re.sub(re.escape(term), "[REDACTED]", redacted, flags=re.IGNORECASE)
    return redacted


def safe_relative(path, root):
    try:
        return str(path.relative_to(root))
    except ValueError:
        return path.name


def slugify(value):
    clean = re.sub(r"[^A-Za-z0-9_.-]+", "-", value).strip("-")
    return clean or "file"


if __name__ == "__main__":
    main()
