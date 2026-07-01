#!/usr/bin/env python3
import json
import sys
from datetime import date, datetime

import openpyxl


def normalize_cell(value):
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return value


def main():
    if len(sys.argv) != 2:
        raise SystemExit("Usage: parse-nsw-contractors.py <Contractor Licence.xlsx>")

    workbook_path = sys.argv[1]
    workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)

    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        rows = worksheet.iter_rows(values_only=True)
        title = next(rows, None)
        headers = next(rows, None)
        if not headers or "Licence Number" not in headers:
            continue

        keys = [str(header).strip() if header is not None else None for header in headers]
        for row in rows:
            record = {
                key: normalize_cell(value)
                for key, value in zip(keys, row)
                if key is not None
            }
            if not record.get("Licence Number"):
                continue
            record["sheetName"] = title[0] if title and title[0] else sheet_name
            print(json.dumps(record, ensure_ascii=False, separators=(",", ":")))


if __name__ == "__main__":
    main()
